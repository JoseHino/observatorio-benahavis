# -*- coding: utf-8 -*-
"""Extracción de la Seguridad Social: afiliación municipal por rama de actividad.

Fichero mensual *Afiliados por Municipios CNAE 2D · Regímenes · Sexo*, con patrón
de URL estable ``https://www.seg-social.es/descargas/STAT/MUNCNAE{MM}{AA}.xlsx``.
Pesa unos 25 MB y contiene medio millón de filas para toda España.

.. warning::
   Tres particularidades verificadas:

   1. **Enmascaramiento.** Los valores entre 1 y 4 se publican como ``<5``. En
      Benahavís esto afecta a cerca de la mitad de las filas, por lo que los
      totales se publican como intervalo (véase :mod:`src.utils.censura`).
   2. **Ruptura de serie.** Hasta diciembre de 2025 la clasificación es CNAE-2009;
      desde enero de 2026 es CNAE-2025. Las series **no se empalman**.
   3. **Ficheros revisados.** Algunos meses se republican con sufijo ``R``
      (``MUNCNAE0925R.xlsx``). Se intenta primero la variante revisada.
"""
from __future__ import annotations

import datetime as dt
import io
from typing import Any

import openpyxl

from ..contexto import CNAE_TURISTICOS, COD_INE, DATA_RAW, RUPTURA_CNAE
from ..utils.censura import CENSURADO, Agregado, parsear
from ..utils.http import FuenteNoDisponible, descargar
from ..utils.log import get_logger

log = get_logger("extract.seg_social")

BASE = "https://www.seg-social.es/descargas/STAT"

COL_MUNICIPIO = 3
COL_REGIMEN = 6
COL_CNAE = 7
COL_CNAE_DESC = 8
COL_SEXO = 10
COL_AFILIADOS = 11


def _descargar_mes(anyo: int, mes: int) -> bytes | None:
    """Descarga el fichero de un mes, probando primero la versión revisada.

    La variante con sufijo ``R`` solo existe para algunos meses, así que se tantea
    con un único intento y sin reintentos: su ausencia es lo normal, no una
    incidencia de red.
    """
    sufijo = f"{mes:02d}{anyo % 100:02d}"
    for nombre, intentos in ((f"MUNCNAE{sufijo}R.xlsx", 1), (f"MUNCNAE{sufijo}.xlsx", 3)):
        try:
            return descargar(f"{BASE}/{nombre}", dir_raw=DATA_RAW, sufijo=".xlsx",
                             timeout=300, reintentos=intentos, guardar=False)
        except FuenteNoDisponible:
            continue
    return None


def _filas_municipio(contenido: bytes) -> list[tuple]:
    """Extrae únicamente las filas de Benahavís del libro de cálculo."""
    libro = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = libro[libro.sheetnames[0]]
    filas = [f for f in hoja.iter_rows(values_only=True)
             if str(f[COL_MUNICIPIO]).strip() == COD_INE]
    libro.close()
    return filas


def afiliacion(meses: list[tuple[int, int]],
               previo: dict[str, Any] | None = None) -> dict[str, Any]:
    """Afiliación a la Seguridad Social en Benahavís por rama de actividad.

    Cada fichero mensual pesa unos 25 MB, así que la descarga es **incremental**:
    los meses que ya figuran en la publicación anterior se reutilizan y solo se
    descargan los nuevos. En la ejecución mensual automática esto reduce la
    descarga de varios cientos de megas a uno o dos ficheros.

    Args:
        meses: lista de ``(año, mes)`` que debe cubrir la serie, en orden cronológico.
        previo: contenido publicado en la ejecución anterior, del que se reutilizan
            los meses ya calculados.

    Returns:
        Series mensuales del total municipal (como intervalo, por la censura), del
        empleo en las cuatro ramas turísticas y del desglose completo por CNAE del
        último mes disponible.
    """
    cache_total = {p["t"]: p for p in (previo or {}).get("serie_total", [])}
    cache_turistico = {p["t"]: p for p in (previo or {}).get("serie_turistico", [])}
    pendientes = [(a, m) for a, m in meses if f"{a}-{m:02d}" not in cache_total]
    log.info("Seguridad Social · afiliación municipal por CNAE · %d meses "
             "(%d reutilizados, %d por descargar)",
             len(meses), len(meses) - len(pendientes), len(pendientes))

    serie_total: list[dict[str, Any]] = []
    serie_turistico: list[dict[str, Any]] = []
    ultimo_desglose: dict[str, Any] = (previo or {}).get("ultimo", {})
    ultimo_mes: str | None = (previo or {}).get("ultimo_mes")
    total_censuradas = 0

    for anyo, mes in meses:
        t = f"{anyo}-{mes:02d}"
        if t in cache_total:
            serie_total.append(cache_total[t])
            if t in cache_turistico:
                serie_turistico.append(cache_turistico[t])
            total_censuradas += int(cache_total[t].get("celdas_censuradas", 0))
            continue
        contenido = _descargar_mes(anyo, mes)
        if contenido is None:
            log.warning("   %s no publicado todavía", t)
            continue
        filas = _filas_municipio(contenido)
        if not filas:
            log.warning("   %s: sin filas para el municipio %s", t, COD_INE)
            continue

        total = Agregado()
        turistico = Agregado()
        por_cnae: dict[str, Agregado] = {}
        etiquetas: dict[str, str] = {}

        for f in filas:
            cnae = str(f[COL_CNAE] or "").strip()
            valor = f[COL_AFILIADOS]
            total.añadir(valor, etiqueta=cnae)
            if cnae in CNAE_TURISTICOS:
                turistico.añadir(valor, etiqueta=cnae)
            if cnae:
                por_cnae.setdefault(cnae, Agregado()).añadir(valor)
                etiquetas.setdefault(cnae, str(f[COL_CNAE_DESC] or "").strip())

        serie_total.append({"t": t, **total.a_dict()})
        serie_turistico.append({"t": t, **turistico.a_dict()})
        total_censuradas += total.censuradas
        if ultimo_mes is None or t >= ultimo_mes:
            ultimo_mes = t
        ultimo_desglose = {
            "mes": t,
            "clasificacion": "CNAE-2025" if t >= RUPTURA_CNAE else "CNAE-2009",
            "ramas": sorted(
                ({"cnae": c, "descripcion": etiquetas.get(c, ""), **a.a_dict()}
                 for c, a in por_cnae.items()),
                key=lambda x: -x["min"],
            ),
            "turisticas": [
                {"cnae": c, "descripcion": CNAE_TURISTICOS[c],
                 **(por_cnae[c].a_dict() if c in por_cnae else Agregado().a_dict())}
                for c in CNAE_TURISTICOS
            ],
        }
        log.info("   %s: %d filas · %d afiliados visibles · %d celdas <5",
                 t, len(filas), total.visible, total.censuradas)

    return {
        "serie_total": serie_total,
        "serie_turistico": serie_turistico,
        "ultimo": ultimo_desglose,
        "ultimo_mes": ultimo_mes,
        "ruptura_cnae": RUPTURA_CNAE,
        "nota_censura": (
            "Los valores entre 1 y 4 se publican como «<5». Los totales se expresan como "
            "intervalo [mínimo observado, mínimo + 4 × número de celdas censuradas]."
        ),
        "celdas_censuradas_acumuladas": total_censuradas,
    }


def meses_recientes(n: int = 18) -> list[tuple[int, int]]:
    """Los ``n`` meses anteriores al actual, en orden cronológico."""
    hoy = dt.date.today()
    salida = []
    anyo, mes = hoy.year, hoy.month
    for _ in range(n):
        mes -= 1
        if mes == 0:
            mes, anyo = 12, anyo - 1
        salida.append((anyo, mes))
    return sorted(salida)
