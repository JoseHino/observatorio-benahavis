# -*- coding: utf-8 -*-
"""Extracción de la Encuesta de Ocupación Hotelera por **zona turística**.

Benahavís no tiene serie municipal de pernoctaciones: la EOH se difunde por zonas
y puntos turísticos, y el municipio no constituye punto turístico propio. La zona
turística **Costa del Sol (Málaga)**, a la que el municipio pertenece, sí se
publica íntegra —pernoctaciones, grado de ocupación, estancia media y plazas
estimadas— con serie mensual desde enero de 2012.

.. important::
   Es un **indicador sustitutivo de ámbito supramunicipal**. Describe el
   comportamiento del alojamiento reglado de toda la Costa del Sol malagueña, no
   el de Benahavís. Todo gráfico que lo use debe declarar ese ámbito, y sus
   cifras no se suman ni se comparan con las de los bloques municipales.

Dos trampas verificadas del endpoint de Dataestur:

1. ``EOH_ZONA_TUR_DL`` devuelve un **fichero XLSX**, no un CSV, pese a que las
   demás operaciones ``*_DL`` del mismo API sirven CSV en ``latin-1``. La
   respuesta llega con ``Content-Type: application/vnd.ms-excel`` y cabecera
   ``PK``: parsearla como texto delimitado produce una sola columna de basura.
2. El endpoint devuelve ``504`` con frecuencia —también en peticiones seguidas
   correctas—, de modo que necesita más reintentos y esperas más largas que el
   resto del pipeline. ``EOH_PUNT_TUR_DL`` (punto turístico) no ha llegado a
   responder en ninguna de las pruebas y no se integra.
"""
from __future__ import annotations

import io
import unicodedata
from typing import Any

import openpyxl

from ..contexto import DATA_RAW
from ..utils.http import descargar
from ..utils.log import get_logger

log = get_logger("extract.eoh")

URL = "https://www.dataestur.es/API-SEGITTUR-v2/EOH_ZONA_TUR_DL"

#: Etiqueta de la zona turística en el fichero: «Costa Del Sol (Málaga), Andalucía».
ZONA = "costa del sol"

#: Fila de agregado de la variable de residencia; las otras dos son sus componentes.
TOTAL_RESIDENCIA = "total"

NOMBRE_ZONA = "Costa del Sol (Málaga)"


def _clave(texto: Any) -> str:
    """Normaliza un rótulo de cabecera: sin tildes, sin espacios y en minúsculas."""
    sin_tildes = unicodedata.normalize("NFKD", str(texto or ""))
    sin_tildes = "".join(c for c in sin_tildes if not unicodedata.combining(c))
    return sin_tildes.strip().lower().replace(" ", "_")


def _numero(valor: Any) -> float | None:
    """Convierte a número; devuelve ``None`` en las celdas vacías del fichero.

    Las filas desglosadas por residencia dejan en blanco el grado de ocupación y
    las plazas estimadas, que solo se publican en la fila de total. Un blanco no
    es un cero y no puede convertirse en uno.
    """
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def _filas_zona() -> list[dict[str, Any]]:
    """Descarga el fichero y devuelve solo las filas de la zona turística."""
    contenido = descargar(URL, dir_raw=DATA_RAW, sufijo=".xlsx", timeout=300,
                          reintentos=5, espera_inicial=15.0)
    libro = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = libro[libro.sheetnames[0]]
    iterador = hoja.iter_rows(values_only=True)
    cabecera = [_clave(c) for c in next(iterador)]

    filas = []
    for cruda in iterador:
        registro = dict(zip(cabecera, cruda))
        if ZONA in str(registro.get("zona_turistica") or "").lower():
            filas.append(registro)
    libro.close()
    if not filas:
        raise ValueError(f"el fichero de la EOH no contiene la zona turística «{NOMBRE_ZONA}»")
    return filas


def zona_turistica() -> dict[str, Any]:
    """Serie de la EOH para la zona turística Costa del Sol (Málaga).

    Returns:
        Serie mensual del total, desglose de pernoctaciones por residencia,
        agregado anual y perfil estacional medio de los años completos recientes.
    """
    log.info("Dataestur · EOH por zona turística (%s)", NOMBRE_ZONA)
    filas = _filas_zona()

    mensual: dict[str, dict[str, Any]] = {}
    pernoctaciones_es: dict[str, float] = {}
    pernoctaciones_ex: dict[str, float] = {}

    for f in filas:
        try:
            anyo, mes = int(f["ano"]), int(f["mes"])
        except (KeyError, TypeError, ValueError):
            continue
        t = f"{anyo}-{mes:02d}"
        residencia = str(f.get("lugar_residencia") or "").lower()
        pernoctaciones = _numero(f.get("pernoctaciones"))

        if TOTAL_RESIDENCIA in residencia:
            mensual[t] = {
                "t": t,
                "viajeros": _numero(f.get("viajeros")),
                "pernoctaciones": pernoctaciones,
                "estancia_media": _numero(f.get("estancia_media")),
                "ocupacion_plazas": _numero(f.get("grado_ocupa_plazas")),
                "ocupacion_fin_semana": _numero(f.get("grado_ocupa_plazas_fin_semana")),
                "plazas_estimadas": _numero(f.get("plazas_estimadas")),
                "establecimientos": _numero(f.get("establecimientos_estimados")),
                "personal_empleado": _numero(f.get("personal_empleado")),
            }
        elif "extranjero" in residencia:
            pernoctaciones_ex[t] = pernoctaciones or 0
        else:
            pernoctaciones_es[t] = pernoctaciones or 0

    meses = sorted(mensual)
    if not meses:
        raise ValueError("la EOH no devolvió ninguna fila de total para la zona turística")

    # — Agregado anual sobre los años naturales completos.
    anual: list[dict[str, Any]] = []
    for anyo in sorted({t[:4] for t in meses}):
        del_anyo = [mensual[t] for t in meses if t.startswith(anyo)]
        if len(del_anyo) < 12:
            continue
        ocupaciones = [m["ocupacion_plazas"] for m in del_anyo if m["ocupacion_plazas"] is not None]
        anual.append({
            "t": anyo,
            "pernoctaciones": sum(m["pernoctaciones"] or 0 for m in del_anyo),
            "viajeros": sum(m["viajeros"] or 0 for m in del_anyo),
            "ocupacion_media": round(sum(ocupaciones) / len(ocupaciones), 2) if ocupaciones else None,
        })

    # — Perfil estacional: media de cada mes calendario en los cinco últimos años
    #   completos. Se excluyen 2020 y 2021 por las restricciones de movilidad, que
    #   deforman cualquier media de estacionalidad.
    anyos_completos = [a["t"] for a in anual if a["t"] not in ("2020", "2021")][-5:]
    estacionalidad = []
    for mes in range(1, 13):
        valores = [mensual[f"{a}-{mes:02d}"]["ocupacion_plazas"]
                   for a in anyos_completos
                   if mensual.get(f"{a}-{mes:02d}", {}).get("ocupacion_plazas") is not None]
        estacionalidad.append({
            "mes": mes,
            "ocupacion_media": round(sum(valores) / len(valores), 2) if valores else None,
        })

    ultimo = mensual[meses[-1]]
    log.info("   %d meses (%s → %s) · último: %s %% de ocupación, %s pernoctaciones",
             len(meses), meses[0], meses[-1], ultimo["ocupacion_plazas"], ultimo["pernoctaciones"])

    return {
        "zona": NOMBRE_ZONA,
        "serie_mensual": [mensual[t] for t in meses],
        "pernoctaciones_por_residencia": {
            "espana": [{"t": t, "v": pernoctaciones_es[t]} for t in sorted(pernoctaciones_es)],
            "extranjero": [{"t": t, "v": pernoctaciones_ex[t]} for t in sorted(pernoctaciones_ex)],
        },
        "anual": anual,
        "estacionalidad": estacionalidad,
        "anyos_estacionalidad": anyos_completos,
        "ultimo": ultimo,
        "ambito": "zona_turistica",
        "es_proxy": True,
        "advertencia": (
            "Indicador sustitutivo de ámbito supramunicipal. Describe el alojamiento reglado "
            "de la zona turística Costa del Sol (Málaga) en su conjunto, no el de Benahavís, "
            "que no constituye punto turístico propio ni dispone de serie municipal de la EOH."
        ),
        "fuente": "INE, Encuesta de Ocupación Hotelera por zonas turísticas (vía Dataestur, SEGITTUR)",
    }
